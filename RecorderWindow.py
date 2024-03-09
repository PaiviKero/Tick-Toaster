import customtkinter as ctk
from defs import *
import Table
import pyscreenshot as ImageGrab
import os
import datetime
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.cell import Cell
from matplotlib import colors

class RecorderWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.times = {key: {} for key in TIME_TYPES}
        self.configure(fg_color=COLOR_BASE_SECONDARY)

        self.geometry(str(RECORDER_WINDOW_WIDTH)+"x"+str(RECORDER_WINDOW_HEIGHT)+"+"+str(RECORDER_WINDOW_START_X)+"+"+str(RECORDER_WINDOW_START_Y))
        self.title('Timer Log')        
        self.tables_frame = ctk.CTkFrame(self, fg_color=COLOR_BASE_SECONDARY)
        self.tables_frame.grid()

    def add_entry(self, name, type, speech_length, time_in_milliseconds, color):
        last_key = 0
        if(len(self.times[type]) > 0):
            last_key = list(self.times[type].keys())[-1]
        next_position = last_key+1
        self.times[type][next_position] = {
            "speaker": name,
            "length": speech_length,
            "time": time_in_milliseconds,
            "color": color
        }
    
    def draw(self):
        self.tables_frame.destroy()
        self.tables_frame = ctk.CTkFrame(self, fg_color=COLOR_BASE_SECONDARY)
        self.tables_frame.grid()
        row_counter = 0
        self.save_button = ctk.CTkButton(master=self.tables_frame, width=100, text="Save manually made changes", command=self.save_changes)
        self.save_button.grid(row=row_counter, column=0, padx=(5,0), pady=(0,5))
        self.export_button = ctk.CTkButton(master=self.tables_frame, text="Export logs to file", command=self.save)
        self.export_button.grid(row=row_counter, column=1, padx=(5,0), pady=(0,5))
        row_counter += 1
        for key in self.times.keys():
            if(key in LOGGED_TIME_TYPES):
                ctk.CTkLabel(self.tables_frame, text=key, font=('Arial',16)).grid(row=row_counter, columnspan=2)
                row_counter += 1
                table_frame = ctk.CTkFrame(self.tables_frame, height=10, fg_color=COLOR_BASE_SECONDARY)
                table_frame.grid(row=row_counter, columnspan=2)
                row_counter += 1
                Table.Table(table_frame, self.times[key], self.delete_entry, key)

    def grab(self):
        return ImageGrab.grab(bbox=(self.winfo_x()+8, self.winfo_y(), self.winfo_width()+58,self.winfo_height()+80))

    def save_changes(self):
        last_label = None
        for widget in self.tables_frame.winfo_children():
            if(isinstance(widget, ctk.CTkLabel)):
                last_label = widget.cget("text")
            if(isinstance(widget, ctk.CTkFrame)):
                counter = 0
                key_index = None
                speaker = None
                speech_length = None
                speech_time = None 
                for child_widget in widget.winfo_children():
                    if(isinstance(child_widget, ctk.CTkEntry)):
                        entry_text = child_widget.get()
                        if(counter == 0):
                            key_index = entry_text
                        if(counter == 1):
                            speaker = entry_text
                        if(counter == 2):
                            speech_length = entry_text
                        if(counter == 3):
                            speech_time = entry_text
                        counter += 1
                    if(isinstance(child_widget, ctk.CTkButton)):
                        counter = 0
                        if(last_label and speaker and key_index):
                            self.times[last_label][int(key_index)]["speaker"] = speaker
                            self.times[last_label][int(key_index)]["length"] = speech_length
                            self.times[last_label][int(key_index)]["time"] = speech_time
                            
    def save(self):
        self.save_image()
        #self.save_csv()
        self.save_xlsx()

    def delete_entry(self, entry_type, entry_key):
        del self.times[entry_type][entry_key]
        self.draw()

    def get_file_name(self):
        date_time_string = datetime.datetime.now().strftime("%B_%d_%Y_%I_%M%p")
        return 'Timer_Report_'+date_time_string

    def save_image(self):
        im = self.grab()
        file_name = self.get_file_name()+'.png'
        im.save(file_name)
        os.startfile(file_name)

    def get_speech_length(self, type):  
        speech_length = ""
        if(type == TIME_TYPES[1]):
            speech_length = str(SPEECH_OPTIONS[TIME_TYPES[1]]["color_times"][0])+"-"+str(SPEECH_OPTIONS[TIME_TYPES[1]]["color_times"][2])+" min"
        if(type == TIME_TYPES[2]):
            speech_length = str(SPEECH_OPTIONS[TIME_TYPES[2]]["color_times"][0])+"-"+str(SPEECH_OPTIONS[TIME_TYPES[2]]["color_times"][2])+" min"
        return speech_length

    def save_csv(self):
        file_name = self.get_file_name()+'.csv'
        with open(file_name, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            field = ["Type", "Speaker", "Length", "Time"]
            
            writer.writerow(field)
            for key in self.times.keys():
                if(key in LOGGED_TIME_TYPES):
                    writer.writerow([key, "", ""])
                    content_row = self.times[key]
                    for key_index in content_row:
                        writer.writerow(["", content_row[key_index]["speaker"], content_row[key_index]["length"], content_row[key_index]["time"]])
        os.startfile(file_name)
    
    def save_xlsx(self):
        wb = Workbook()
        worksheet = wb.active

        worksheet.append(["Type", "Speaker", "Length", "Time"]);
        for key in self.times.keys():
            if(key in LOGGED_TIME_TYPES):
                worksheet.append([key, "", ""])
                content_row = self.times[key]
                for key_index in content_row:
                    time_cell = Cell(worksheet, value=content_row[key_index]["time"])
                    hex_color = colors.cnames[content_row[key_index]["color"].lower().replace(" ", "")].replace("#", "FF")
                    time_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type = "solid")
                    worksheet.append(["", content_row[key_index]["speaker"], content_row[key_index]["length"], time_cell])

        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value

        file_name = self.get_file_name()+'.xlsx'
        wb.save(file_name)
        os.startfile(file_name)
    
